"""Export the canonical WHestBench research ledger to browsable CSVs.

The ledger workbook is a local research artifact and is not tracked. This
script converts it into per-sheet CSVs under ``whestbench/ledger/`` so that
reviewers can read the full experiment record without Excel, and so that the
evidence-status columns survive review in plain text.

Usage::

    python scripts/export_ledger.py path/to/ledger.xlsx

With no argument it looks for the v31 reconciled ledger in the usual local
location.
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "whestbench" / "ledger"
DEFAULT_LEDGER = Path.home() / (
    "Downloads/Whestbench/"
    "whestbench_canonical_research_ledger_20260731_reconciled_v31_final_local_writeup.xlsx"
)


def slugify(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
    return slug or "sheet"


def export(workbook_path: Path) -> list[tuple[str, str, int, int]]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    exported: list[tuple[str, str, int, int]] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = [
            ["" if cell is None else str(cell).replace("\r\n", "\n").strip() for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]
        while rows and not any(cell for cell in rows[-1]):
            rows.pop()
        if not rows:
            continue
        width = max(len(row) for row in rows)
        rows = [row + [""] * (width - len(row)) for row in rows]

        slug = slugify(sheet_name)
        target = OUT_DIR / f"{slug}.csv"
        with target.open("w", newline="", encoding="utf-8") as stream:
            csv.writer(stream).writerows(rows)
        exported.append((sheet_name, target.name, len(rows), width))

    return exported


def write_index(exported: list[tuple[str, str, int, int]], workbook_path: Path) -> None:
    lines = [
        "# Canonical research ledger",
        "",
        "Every sheet of the canonical WHestBench research ledger, exported to CSV",
        "so it can be read, diffed, and searched without Excel. This is the full",
        "internal experiment record, including entries that were later retracted,",
        "quarantined, or reclassified. Read the evidence-status columns before",
        "quoting any row.",
        "",
        f"Source workbook: `{workbook_path.name}`",
        "",
        "Regenerate with:",
        "",
        "```bash",
        "python scripts/export_ledger.py path/to/ledger.xlsx",
        "```",
        "",
        "| sheet | file | rows | cols |",
        "|---|---|---:|---:|",
    ]
    for sheet_name, filename, n_rows, n_cols in exported:
        lines.append(f"| {sheet_name} | [`{filename}`]({filename}) | {n_rows} | {n_cols} |")
    lines += [
        "",
        "## How to read this",
        "",
        "The ledger is a working record, not a results table. It was maintained",
        "across many agent sessions and repeatedly reconciled; the `Reconciliation",
        "Audit`, `Contradiction Map`, and `Evidence Quarantine` sheets exist",
        "precisely because earlier rows disagreed with later measurement. A row",
        "appearing here is not a claim that its number is correct.",
        "",
        "For claims that survived review with their evidence boundary attached,",
        "use [`../claims.csv`](../claims.csv). For the graded competition result,",
        "use [`../phase1_320802.json`](../phase1_320802.json).",
        "",
    ]
    (OUT_DIR / "README.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_LEDGER
    if not workbook_path.is_file():
        print(f"ledger workbook not found: {workbook_path}", file=sys.stderr)
        return 1
    exported = export(workbook_path)
    write_index(exported, workbook_path)
    total = sum(rows for _, _, rows, _ in exported)
    print(f"exported {len(exported)} sheets, {total} rows -> {OUT_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
